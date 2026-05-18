from io import BytesIO

from openpyxl import Workbook


def build_export_xlsx(payload: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "result"
    items = payload.get("items", [])
    supplier_names = payload.get("supplier_names", {})

    supplier_ids: list[int] = [int(key) for key in supplier_names.keys()]
    if not supplier_ids:
        for item in items:
            for match in item.get("matches", []):
                supplier_id = int(match["supplier_id"])
                if supplier_id not in supplier_ids:
                    supplier_ids.append(supplier_id)

    header = ["Продукт", "Ед.", "Кол-во к заказу"]
    for supplier_id in supplier_ids:
        name = supplier_names.get(str(supplier_id), f"S{supplier_id}")
        header.append(f"Цена {name}")
    for supplier_id in supplier_ids:
        name = supplier_names.get(str(supplier_id), f"S{supplier_id}")
        header.append(f"Кол-во {name}")
    for supplier_id in supplier_ids:
        name = supplier_names.get(str(supplier_id), f"S{supplier_id}")
        header.append(f"Сумма {name}")
    header.append("Комментарий")
    ws.append(header)

    price_start = 4
    qty_start = price_start + len(supplier_ids)
    amount_start = qty_start + len(supplier_ids)

    # First data row is total row by requirement.
    total_row_idx = 2
    ws.append(["Итого", "", len(items)] + [0] * (len(header) - 3))

    for row_idx, item in enumerate(items, start=3):
        row = [item.get("canonical_name", ""), item.get("unit", ""), float(item.get("quantity", 0))]
        matches_by_supplier = {int(m["supplier_id"]): m for m in item.get("matches", [])}
        alloc_by_supplier = {int(a["supplier_id"]): a for a in item.get("allocation", [])}
        for supplier_id in supplier_ids:
            price = float(matches_by_supplier.get(supplier_id, {}).get("price", 0))
            row.append(price)
        for supplier_id in supplier_ids:
            alloc_qty = float(alloc_by_supplier.get(supplier_id, {}).get("quantity", 0))
            row.append(alloc_qty)
        for _ in supplier_ids:
            row.append(None)
        row.append(item.get("comment", ""))
        ws.append(row)

        for idx, _ in enumerate(supplier_ids):
            price_col = _xl_col(price_start + idx)
            qty_col = _xl_col(qty_start + idx)
            amount_col = _xl_col(amount_start + idx)
            ws[f"{amount_col}{row_idx}"] = f"={price_col}{row_idx}*{qty_col}{row_idx}"

    last_item_row = len(items) + 2
    for idx in range(len(supplier_ids)):
        amount_col = _xl_col(amount_start + idx)
        ws[f"{amount_col}{total_row_idx}"] = f"=SUM({amount_col}3:{amount_col}{last_item_row})"

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _xl_col(index: int) -> str:
    result = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result
