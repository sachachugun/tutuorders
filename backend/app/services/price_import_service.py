from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
import xlrd
from fastapi import HTTPException
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.models import CanonicalProduct, Price, Supplier, SupplierSku

MAX_ROWS = 500
ALLOWED_UNITS = {"кг", "г", "л", "мл"}


def _to_float(raw_value: object) -> float:
    if raw_value is None:
        return 0
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    text = (
        str(raw_value)
        .strip()
        .replace("\xa0", "")
        .replace(" ", "")
        .replace(",", ".")
    )
    return float(text) if text else 0


def _normalize_unit(raw_unit: object) -> str:
    unit = str(raw_unit or "").strip().lower()
    if unit == "гр":
        unit = "г"
    return unit if unit in ALLOWED_UNITS else "кг"


def parse_price_file(file_name: str, file_content: bytes, supplier_id: int) -> list[dict]:
    if not file_content:
        raise HTTPException(status_code=400, detail="Файл пустой. Загрузите прайс в формате .xls или .xlsx")
    ext = (file_name or "").lower()
    if ext.endswith(".xlsx"):
        return _parse_xlsx(file_content)
    if ext.endswith(".xls"):
        return _parse_xls(file_content)
    raise HTTPException(status_code=400, detail="Поддерживаются только .xls и .xlsx")


def _parse_xlsx(file_content: bytes) -> list[dict]:
    workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.active
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    has_header = _looks_like_header(header)
    start_row = 2 if has_header else 1
    rows: list[dict] = []
    for index, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=1):
        if index > MAX_ROWS:
            raise HTTPException(status_code=400, detail="Лимит прайса: не более 500 строк")
        rows.append(_extract_row(row, 0, 1, 2))
    if not rows:
        raise HTTPException(status_code=400, detail="Не найдено строк с данными в прайсе")
    return rows


def _parse_xls(file_content: bytes) -> list[dict]:
    workbook = xlrd.open_workbook(file_contents=file_content)
    sheet = workbook.sheet_by_index(0)
    header = sheet.row_values(0) if sheet.nrows > 0 else []
    has_header = _looks_like_header(header)
    start_row = 1 if has_header else 0
    rows: list[dict] = []
    for index, row_idx in enumerate(range(start_row, sheet.nrows), start=1):
        if index > MAX_ROWS:
            raise HTTPException(status_code=400, detail="Лимит прайса: не более 500 строк")
        row = sheet.row_values(row_idx)
        rows.append(_extract_row(row, 0, 1, 2))
    if not rows:
        raise HTTPException(status_code=400, detail="Не найдено строк с данными в прайсе")
    return rows


def _extract_row(row: object, col_name: int, col_unit: int, col_price: int) -> dict:
    name_in_price = str((row[col_name] if len(row) > col_name else "") or "").strip()
    raw_unit = str((row[col_unit] if len(row) > col_unit else "") or "").strip()
    unit = _normalize_unit(raw_unit)
    price = _to_float(row[col_price] if len(row) > col_price else 0)
    return {"name_in_price": name_in_price, "unit": unit, "raw_unit": raw_unit, "price": price}


def _looks_like_header(header: object) -> bool:
    normalized = [str(h or "").strip().lower() for h in (header or [])]
    if len(normalized) < 3:
        return False
    col_a = normalized[0]
    col_b = normalized[1]
    col_c = normalized[2]
    return (
        any(token in col_a for token in ("номенклат", "товар", "продукт", "наименование", "name"))
        and any(token in col_b for token in ("ед", "unit", "единиц"))
        and any(token in col_c for token in ("цена", "price", "стоим"))
    )


def normalize_price_rows(rows: list[dict]) -> tuple[list[dict], dict]:
    dedup = defaultdict(lambda: None)
    stats = {"invalid_price": 0, "empty_name": 0, "duplicates_removed": 0, "section_rows": 0}
    for row in rows:
        name = (row.get("name_in_price") or "").strip()
        if not name:
            stats["empty_name"] += 1
            continue
        price = float(row.get("price", 0))
        if price <= 0:
            if _is_section_row(name, str(row.get("raw_unit", ""))):
                stats["section_rows"] += 1
            else:
                stats["invalid_price"] += 1
            continue

        existing = dedup[name]
        if existing is None or price < existing["price"]:
            if existing is not None:
                stats["duplicates_removed"] += 1
            dedup[name] = {"name_in_price": name, "unit": row.get("unit", "кг"), "price": round(price, 2)}
        else:
            stats["duplicates_removed"] += 1
    normalized = list(dedup.values())
    if not normalized:
        raise HTTPException(
            status_code=400,
            detail=(
                "Не найдено валидных позиций с ценой > 0. "
                "Проверьте, что в файле заполнены название, единица и цена."
            ),
        )
    return normalized, stats


def _price_name_key(name: str) -> str:
    """Ключ для сопоставления: без лишних пробелов, без учёта регистра."""
    return " ".join(str(name or "").strip().split()).casefold()


def _build_prices_lookup(prices: list[Price]) -> tuple[dict[str, Price], dict[str, Price]]:
    exact: dict[str, Price] = {}
    normalized: dict[str, Price] = {}
    for row in prices:
        exact[row.name_in_price] = row
        norm = _price_name_key(row.name_in_price)
        if norm not in normalized:
            normalized[norm] = row
    return exact, normalized


def _find_price_for_sku_name(
    sku_name: str,
    exact: dict[str, Price],
    normalized: dict[str, Price],
) -> Price | None:
    if not sku_name:
        return None
    hit = exact.get(sku_name)
    if hit:
        return hit
    return normalized.get(_price_name_key(sku_name))


def relink_supplier_skus_from_prices(db: Session, supplier_id: int) -> dict:
    """Привязать SKU к строкам прайса по названию (точное или нормализованное совпадение)."""
    prices = db.query(Price).filter(Price.supplier_id == supplier_id).all()
    exact, normalized = _build_prices_lookup(prices)
    skus = db.query(SupplierSku).filter(SupplierSku.supplier_id == supplier_id).all()
    relinked = 0
    broken = 0
    broken_items: list[dict] = []
    affected_product_ids: set[int] = set()
    for sku in skus:
        matched = _find_price_for_sku_name(sku.name_in_price, exact, normalized)
        if matched:
            sku.price_id = matched.id
            sku.unit = matched.unit
            sku.price = float(matched.price)
            relinked += 1
        else:
            sku.price_id = None
            if sku.name_in_price:
                broken += 1
                pid = int(sku.canonical_product_id)
                affected_product_ids.add(pid)
                product = db.get(CanonicalProduct, pid)
                broken_items.append(
                    {
                        "product_id": pid,
                        "product_name": product.name if product else f"#{pid}",
                        "name_in_price": sku.name_in_price,
                    }
                )
    broken_items.sort(key=lambda row: (row["product_name"].lower(), row["name_in_price"].lower()))
    return {
        "relinked_count": relinked,
        "broken_sku_links_count": broken,
        "affected_products_count": len(affected_product_ids),
        "broken_items": broken_items,
        "price_rows_count": len(prices),
    }


def _is_section_row(name: str, raw_unit: str) -> bool:
    if raw_unit.strip():
        return False
    normalized = name.strip()
    if not normalized:
        return False
    # Typical section rows: short uppercase category labels like "ГРИБЫ", "ОВОЩИ".
    return len(normalized.split()) <= 3 and normalized.upper() == normalized


def replace_supplier_prices(db: Session, supplier_id: int, normalized_rows: list[dict]) -> dict:
    old_prices = db.query(Price).filter(Price.supplier_id == supplier_id).all()
    old_names = {row.name_in_price for row in old_prices}

    db.execute(delete(Price).where(Price.supplier_id == supplier_id))
    for row in normalized_rows:
        db.add(
            Price(
                supplier_id=supplier_id,
                name_in_price=row["name_in_price"],
                unit=row["unit"],
                price=row["price"],
            )
        )
    # autoflush=False в Session — без flush новые prices не видны в query, все SKU обнулялись.
    db.flush()

    supplier = db.get(Supplier, supplier_id)
    if supplier:
        supplier.last_price_upload_at = datetime.now(timezone.utc)

    prices = db.query(Price).filter(Price.supplier_id == supplier_id).all()
    new_names = {row.name_in_price for row in prices}
    new_price_items_count = len(new_names - old_names)

    relink_stats = relink_supplier_skus_from_prices(db, supplier_id)

    db.commit()
    return {
        "new_price_items_count": new_price_items_count,
        "broken_sku_links_count": relink_stats["broken_sku_links_count"],
        "affected_products_count": relink_stats["affected_products_count"],
        "relinked_count": relink_stats["relinked_count"],
        "broken_items": relink_stats["broken_items"],
    }
