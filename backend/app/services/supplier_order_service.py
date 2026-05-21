import re

from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models import (
    Allocation,
    DemandLine,
    Department,
    Location,
    ProcurementBatch,
    Supplier,
    SupplierOrderLine,
    SupplierSku,
)
from app.services.spec_service import resolve_spec_text


def _line_comment(spec_text: str, line_notes: str | None) -> str:
    parts: list[str] = []
    if spec_text.strip():
        parts.append(spec_text.strip())
    if line_notes and line_notes.strip():
        parts.append(line_notes.strip())
    return "\n".join(parts)


def build_supplier_orders(db: Session, batch_id: int) -> dict:
    batch = db.get(ProcurementBatch, batch_id)
    if not batch:
        raise ValueError("batch_not_found")

    allocations = db.scalars(select(Allocation).where(Allocation.batch_id == batch_id)).all()
    if not allocations:
        raise ValueError("no_allocations")

    db.execute(delete(SupplierOrderLine).where(SupplierOrderLine.batch_id == batch_id))

    sort_order = 0
    for alloc in allocations:
        line = db.get(DemandLine, alloc.demand_line_id)
        sku = db.get(SupplierSku, alloc.supplier_sku_id)
        if not line or not sku:
            continue

        spec_text = ""
        if line.canonical_product_id:
            spec = resolve_spec_text(
                db,
                int(line.canonical_product_id),
                supplier_id=alloc.supplier_id,
                location_id=line.location_id,
                department_id=line.department_id,
            )
            spec_text = spec.get("spec_text") or ""
        comment = _line_comment(spec_text, line.line_notes)

        db.add(
            SupplierOrderLine(
                batch_id=batch_id,
                supplier_id=alloc.supplier_id,
                location_id=line.location_id,
                department_id=line.department_id,
                allocation_id=alloc.id,
                supplier_product_name=sku.name_in_price,
                quantity=float(alloc.quantity),
                unit=alloc.unit,
                unit_price=float(alloc.unit_price),
                amount=float(alloc.amount),
                spec_text=spec_text or None,
                line_comment=comment or None,
                sort_order=sort_order,
            )
        )
        sort_order += 1

    if batch.status == "optimized":
        from app.services.procurement_batch_meta import apply_batch_status

        apply_batch_status(batch, "approved")
    db.commit()
    return get_supplier_orders_state(db, batch_id)


def _order_line_to_dict(db: Session, row: SupplierOrderLine) -> dict:
    location = db.get(Location, row.location_id)
    department = db.get(Department, row.department_id)
    supplier = db.get(Supplier, row.supplier_id)
    return {
        "id": row.id,
        "batch_id": row.batch_id,
        "supplier_id": row.supplier_id,
        "supplier_name": supplier.name if supplier else "",
        "location_id": row.location_id,
        "location_name": location.name if location else "",
        "department_id": row.department_id,
        "department_name": department.name if department else "",
        "allocation_id": row.allocation_id,
        "supplier_product_name": row.supplier_product_name,
        "quantity": float(row.quantity),
        "unit": row.unit,
        "unit_price": float(row.unit_price),
        "amount": float(row.amount),
        "spec_text": row.spec_text,
        "line_comment": row.line_comment,
        "sort_order": int(row.sort_order),
    }


def get_supplier_orders_state(db: Session, batch_id: int) -> dict:
    batch = db.get(ProcurementBatch, batch_id)
    if not batch:
        raise ValueError("batch_not_found")

    rows = db.scalars(
        select(SupplierOrderLine)
        .where(SupplierOrderLine.batch_id == batch_id)
        .order_by(
            SupplierOrderLine.supplier_id,
            SupplierOrderLine.location_id,
            SupplierOrderLine.department_id,
            SupplierOrderLine.sort_order,
            SupplierOrderLine.id,
        )
    ).all()

    line_dicts = [_order_line_to_dict(db, row) for row in rows]
    groups_map: dict[tuple[int, int, int], dict] = {}
    for line in line_dicts:
        key = (line["supplier_id"], line["location_id"], line["department_id"])
        if key not in groups_map:
            groups_map[key] = {
                "supplier_id": line["supplier_id"],
                "supplier_name": line["supplier_name"],
                "location_id": line["location_id"],
                "location_name": line["location_name"],
                "department_id": line["department_id"],
                "department_name": line["department_name"],
                "lines": [],
                "total_amount": 0.0,
            }
        groups_map[key]["lines"].append(line)
        groups_map[key]["total_amount"] += line["amount"]

    groups = list(groups_map.values())
    for group in groups:
        group["total_amount"] = round(group["total_amount"], 2)

    suppliers_seen: dict[int, str] = {}
    locations_seen: dict[int, str] = {}
    departments_seen: dict[int, str] = {}
    for line in line_dicts:
        suppliers_seen[line["supplier_id"]] = line["supplier_name"]
        locations_seen[line["location_id"]] = line["location_name"]
        departments_seen[line["department_id"]] = line["department_name"]

    created_at = batch.created_at.isoformat() if hasattr(batch.created_at, "isoformat") else str(batch.created_at)

    return {
        "batch_id": batch_id,
        "batch_title": batch.title,
        "batch_status": batch.status,
        "created_at": created_at,
        "lines_count": len(line_dicts),
        "groups_count": len(groups),
        "groups": groups,
        "lines": line_dicts,
        "suppliers": [{"id": k, "name": v} for k, v in sorted(suppliers_seen.items())],
        "locations": [{"id": k, "name": v} for k, v in sorted(locations_seen.items())],
        "departments": [{"id": k, "name": v} for k, v in sorted(departments_seen.items())],
    }


def update_order_line_comment(db: Session, batch_id: int, line_id: int, line_comment: str) -> dict:
    batch = db.get(ProcurementBatch, batch_id)
    if not batch:
        raise ValueError("batch_not_found")
    row = db.get(SupplierOrderLine, line_id)
    if not row or row.batch_id != batch_id:
        raise ValueError("line_not_found")
    row.line_comment = line_comment.strip() or None
    db.commit()
    return _order_line_to_dict(db, row)


def _sanitize_sheet_name(value: str) -> str:
    text = re.sub(r'[:\\/?*\[\]]', "_", (value or "").strip())
    return (text or "sheet")[:31]


def build_procurement_batch_xlsx(db: Session, batch_id: int) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    state = get_supplier_orders_state(db, batch_id)
    if not state["lines_count"]:
        raise ValueError("orders_not_built")

    batch = db.get(ProcurementBatch, batch_id)
    wb = Workbook()
    wb.remove(wb.active)

    for group in state["groups"]:
        sheet_title = _sanitize_sheet_name(
            f"{group['supplier_name']}_{group['location_name']}_{group['department_name']}"
        )
        ws = wb.create_sheet(title=sheet_title)
        ws.append(["План закупки", state["batch_title"]])
        ws.append(["Локация", group["location_name"]])
        ws.append(["Отдел", group["department_name"]])
        ws.append(["Поставщик", group["supplier_name"]])
        ws.append(["Дата", state.get("created_at", "")])
        ws.append([])
        ws.append(["Название", "Кол-во", "Ед.", "Цена", "Сумма", "Комментарий"])

        total = 0.0
        for line in group["lines"]:
            ws.append(
                [
                    line["supplier_product_name"],
                    line["quantity"],
                    line["unit"],
                    line["unit_price"],
                    line["amount"],
                    line["line_comment"] or "",
                ]
            )
            total += line["amount"]
        ws.append([])
        ws.append(["Итого", "", "", "", round(total, 2), ""])

    if not wb.sheetnames:
        ws = wb.create_sheet("empty")
        ws.append(["Нет строк заказа"])

    if batch:
        from app.services.procurement_batch_meta import apply_batch_status

        apply_batch_status(batch, "exported")
        db.commit()

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
